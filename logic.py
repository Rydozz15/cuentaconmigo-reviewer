import pandas as pd
import numpy as np
import openpyxl
import os
import shutil
from datetime import datetime, timedelta

# ==============================================================================
# 🛠️ CONSTANTES Y CONFIGURACIONES
# ==============================================================================
status_map = {
    'Rechazo / Sin disponibilidad': '1. Rechazo / Sin disponibilidad',
    'Recién Notificado (< 5 días)': '2. Recién Notificados',
    'Notificado - No Descarga App': '3. Notificados No Descargan',
    'Notificado con Error Técnico': '4. Notificados con Error Técnico',
    'En Progreso - Atrasado': '5. En Progreso Atrasados',
    'En Libro 1': '6. En Libro 1',
    'En Progreso - Empezando Libro 2': '7. Empezando Libro 2',
    'En Progreso - Próximo a Terminar (Avanzado)': '8. Próximos a Terminar (Avanzados)',
    'Completado': '9. Completados'
}

GLOSARIO_PANTALLAS = {
    'L1B1S0_15': 'Mapa de misiones libro 1',
    'L2B1S0_14': 'Retomar sesión (2)',
    'L1B1S0_16': 'Retomar sesión sin pregunta',
    'L1B2S1_2L21': 'Final primera lectura Adentro',
    'L1B2S1_3': 'Inicio primera actividad Adentro',
    'L1B3S1_33': 'Final primera actividad Adentro',
    'L1B2S2_2L21': 'Final segunda lectura Adentro',
    'L1B2S2_3': 'Inicio segunda actividad Adentro',
    'L1B3S2_9': 'Termina segunda actividad Adentro',
    'L1B2S3_2L21': 'Final tercera lectura Adentro',
    'L1B2S3_3': 'Inicio tercera actividad Adentro',
    'L1B3S3_15': 'Termina tercera actividad Adentro',
    'L1B3S4_14': 'Termina actividad recontado Adentro',
    'L1B3S4_15': 'Encuesta final Adentro',
    'L2B2S1_2L21': 'Final primera lectura Nibaldo',
    'L2B2S1_3': 'Inicio primera actividad Nibaldo',
    'L2B3S1_33': 'Final primera actividad Nibaldo',
    'L2B2S2_2L21': 'Final segunda lectura Nibaldo',
    'L2B2S2_3': 'Inicio segunda actividad Nibaldo',
    'L2B3S2_9': 'Termina segunda actividad Nibaldo',
    'L2B2S3_2L21': 'Final tercera lectura Nibaldo',
    'L2B2S3_3': 'Inicio tercera actividad Nibaldo',
    'L2B3S3_15': 'Termina tercera actividad Nibaldo',
    'L2B3S4_14': 'Termina actividad recontado Nibaldo',
    'L2B3S4_15': 'Encuesta final Nibaldo'
}

PANTALLAS_FINALES = ['L2B3S4_15']

SECUENCIA_PANTALLAS = [
    'L1B1S0', 'L1B2S1', 'L1B3S1', 'L1B2S2', 'L1B3S2', 'L1B2S3', 'L1B3S3', 'L1B3S4',
    'L2B1S0', 'L2B2S1', 'L2B3S1', 'L2B2S2', 'L2B3S2', 'L2B2S3', 'L2B3S3', 'L2B3S4'
]

ETIQUETAS_SECCION = {
    'L1B1S0': 'L1 - Mapa misiones',
    'L1B2S1': 'L1 - Lectura 1 Adentro',
    'L1B3S1': 'L1 - Actividad 1 Adentro',
    'L1B2S2': 'L1 - Lectura 2 Adentro',
    'L1B3S2': 'L1 - Actividad 2 Adentro',
    'L1B2S3': 'L1 - Lectura 3 Adentro',
    'L1B3S3': 'L1 - Actividad 3 Adentro',
    'L1B3S4': 'L1 - Recontado Adentro',
    'L2B1S0': 'L2 - Mapa misiones',
    'L2B2S1': 'L2 - Lectura 1 Nibaldo',
    'L2B3S1': 'L2 - Actividad 1 Nibaldo',
    'L2B2S2': 'L2 - Lectura 2 Nibaldo',
    'L2B3S2': 'L2 - Actividad 2 Nibaldo',
    'L2B2S3': 'L2 - Lectura 3 Nibaldo',
    'L2B3S3': 'L2 - Actividad 3 Nibaldo',
    'L2B3S4': 'L2 - Recontado Nibaldo'
}

# ==============================================================================
# 🔧 FUNCIONES AUXILIARES
# ==============================================================================
def limpiar_tel(val):
    if pd.isna(val): return ""
    s = str(val).replace('.0', '')
    digits = "".join(filter(str.isdigit, s))
    return "56" + digits[-9:] if len(digits) >= 9 else digits

def solo_9_digitos(val):
    if pd.isna(val): return ""
    s = str(val).replace('.0', '')
    digits = "".join(filter(str.isdigit, s))
    return digits[-9:] if len(digits) >= 9 else ""

def parsear_timestamp(val):
    if pd.isna(val) or val == "":
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        if pd.isna(val):
            return None
        return pd.to_datetime(val)
    
    val_str = str(val).strip()
    if val_str.lower() in ['nat', 'nan', 'none', '']:
        return None
        
    formatos = [
        '%Y-%m-%dT%H:%M:%S.%f', # ISO 8601 con T y microsegundos
        '%Y-%m-%dT%H:%M:%S',    # ISO 8601 con T sin microsegundos
        '%Y-%m-%d %H:%M:%S:%f',
        '%Y-%m-%d %H:%M:%S',
        '%d-%m-%Y %H:%M:%S',
        '%d/%m/%Y %H:%M:%S',
        '%Y-%m-%d',
        '%d-%m-%Y',
        '%d/%m/%Y'
    ]
    for fmt in formatos:
        try:
            res = pd.to_datetime(val_str, format=fmt)
            if pd.isna(res):
                return None
            return res
        except (ValueError, TypeError):
            continue
            
    res = pd.to_datetime(val_str, dayfirst=True, errors='coerce')
    if pd.isna(res):
        return None
    return res

def parsear_fecha_inscripcion(val):
    dt = parsear_timestamp(val)
    if dt is None:
        return None
    # Si el día y el mes son ambos <= 12, se asume que Excel los intercambió.
    # Los intercambiamos de vuelta.
    if dt.day <= 12 and dt.month <= 12:
        try:
            return datetime(year=dt.year, month=dt.day, day=dt.month, hour=dt.hour, minute=dt.minute, second=dt.second)
        except ValueError:
            return dt
    return dt

def extraer_seccion(pantalla):
    if pd.isna(pantalla): return None
    s = str(pantalla)
    if '_' in s:
        return s.split('_')[0]
    return s[:6] if len(s) >= 6 else s

def calcular_progreso(pantalla):
    seccion = extraer_seccion(pantalla)
    if seccion is None: return 0
    try:
        idx = SECUENCIA_PANTALLAS.index(seccion)
        return (idx + 1) / len(SECUENCIA_PANTALLAS) * 100
    except ValueError:
        return 0

def valores_diferentes(v1, v2):
    v1_isna = pd.isna(v1) or v1 == "" or str(v1).strip().lower() in ['nat', 'nan', 'none']
    v2_isna = pd.isna(v2) or v2 == "" or str(v2).strip().lower() in ['nat', 'nan', 'none']
    if v1_isna and v2_isna:
        return False
    if v1_isna != v2_isna:
        return True
    try:
        if float(v1) == float(v2):
            return False
    except (ValueError, TypeError):
        pass
        
    try:
        b1 = bool(v1) if str(v1).lower() in ['true', '1', '1.0'] else False
        b2 = bool(v2) if str(v2).lower() in ['true', '1', '1.0'] else False
        if b1 == b2 and str(v1).lower() in ['true', 'false', '1', '0', '1.0', '0.0'] and str(v2).lower() in ['true', 'false', '1', '0', '1.0', '0.0']:
            return False
    except (ValueError, TypeError):
        pass
        
    s1 = str(v1).strip().lower().replace('.0', '')
    s2 = str(v2).strip().lower().replace('.0', '')
    
    d1 = parsear_timestamp(v1)
    d2 = parsear_timestamp(v2)
    if d1 is not None and d2 is not None:
        return d1.strftime('%Y-%m-%d %H:%M:%S') != d2.strftime('%Y-%m-%d %H:%M:%S')
        
    return s1 != s2

# ==============================================================================
# 🚀 PROCESAMIENTO PRINCIPAL
# ==============================================================================
def procesar_datos_completos(df_participantes, df_contactados, df_servidor):
    """
    Cruce de logs y lógica de negocio. Retorna:
    - df_participantes: El DataFrame mapeado e integrado.
    - nuevos_pendientes: Lista de dicts con participantes listos para agregar de CONTACTADOS.
    """
    hoy_dt = datetime.now()
    
    # Asegurar que todas las columnas de df_participantes sean de tipo object para evitar errores de casteo (dtype datetime64)
    df_participantes = df_participantes.copy()
    for col in df_participantes.columns:
        df_participantes[col] = df_participantes[col].astype(object)
    
    # 1. Detectar Nuevos Pendientes en CONTACTADOS
    nuevos_pendientes = []
    IDX_ID = 0
    IDX_ESTADO = 24
    
    try:
        # Columna 24 es el Estado en CONTACTADOS
        aptos = df_contactados[df_contactados.iloc[:, IDX_ESTADO].astype(str).str.strip().str.lower() == 'participante']
        actuales_ids = set(df_participantes.iloc[:, 0].astype(str).str.strip().replace(r'\.0', '', regex=True))
        
        for idx_apt, row_apt in aptos.iterrows():
            pid = str(row_apt.iloc[0]).strip().replace('.0', '')
            if pid not in actuales_ids:
                nuevos_pendientes.append({
                    "id": pid,
                    "colegio": str(row_apt.iloc[1]),
                    "nombre_adulto": f"{str(row_apt.iloc[2])} {str(row_apt.iloc[3])} {str(row_apt.iloc[4])}".strip(),
                    "telefono": str(row_apt.iloc[10]),
                    "nombre_nino": f"{str(row_apt.iloc[5])} {str(row_apt.iloc[6])} {str(row_apt.iloc[7])}".strip()
                })
    except Exception as e:
        print(f"⚠️ Error al cruzar nuevos de CONTACTADOS: {e}")
        
    # 2. Filtrar ruido en reporte
    COL_REGISTRO = 'Registro'
    PATRONES_RUIDO = [
        r'env[ií]o de notificaci[oó]n por inactividad',
        r'notificaci[oó]n de inactividad',
        r'recordatorio autom[aá]tico',
        r'push notification',
        r'system alert'
    ]
    
    if COL_REGISTRO in df_servidor.columns:
        regex_ruido = '|'.join(PATRONES_RUIDO)
        mask_uso_real = ~df_servidor[COL_REGISTRO].astype(str).str.contains(
            regex_ruido, case=False, na=False, regex=True
        )
        df_servidor = df_servidor[mask_uso_real].copy()
        
    df_participantes['tel_match'] = df_participantes['Teléfono'].apply(solo_9_digitos)
    df_servidor['tel_match'] = df_servidor['Numero Telefono'].apply(solo_9_digitos)
    df_servidor['timestamp'] = df_servidor['HoradeEntradaPantalla'].apply(parsear_timestamp)
    
    df_servidor_valido = df_servidor[
        (df_servidor['tel_match'].str.len() == 9) &
        (df_servidor['timestamp'].notna())
    ].copy()
    
    primeros_logs = df_servidor_valido.sort_values('timestamp').groupby('tel_match')['timestamp'].first()
    
    # Progreso por pantalla secuencial
    df_servidor_valido['seccion_tmp'] = df_servidor_valido['IDPantalla'].apply(extraer_seccion)
    df_servidor_valido['seq_idx_tmp'] = df_servidor_valido['seccion_tmp'].apply(
        lambda s: SECUENCIA_PANTALLAS.index(s) if s in SECUENCIA_PANTALLAS else -1
    )
    
    df_valid_seq = df_servidor_valido[df_servidor_valido['seq_idx_tmp'] >= 0]
    if not df_valid_seq.empty:
        furthest_pantallas = df_valid_seq.sort_values(['seq_idx_tmp', 'timestamp']).groupby('tel_match')['IDPantalla'].last()
    else:
        furthest_pantallas = pd.Series(dtype=object)
        
    last_pantallas_all = df_servidor_valido.sort_values('timestamp').groupby('tel_match')['IDPantalla'].last()
    ultimas_pantallas = furthest_pantallas.reindex(last_pantallas_all.index).fillna(last_pantallas_all)
    ultimas_fechas = df_servidor_valido.sort_values('timestamp').groupby('tel_match')['timestamp'].last()
    
    # Completados
    tels_con_final = set(df_servidor_valido[df_servidor_valido['IDPantalla'].isin(PANTALLAS_FINALES)]['tel_match'].unique())
    fechas_final = df_servidor_valido[df_servidor_valido['IDPantalla'].isin(PANTALLAS_FINALES)].sort_values('timestamp').groupby('tel_match')['timestamp'].first()
    
    match_fechas = df_participantes['tel_match'].map(primeros_logs)
    
    df_participantes['Fecha log en app'] = df_participantes['Fecha log en app'].astype(object)
    match_fechas_str = match_fechas.dt.strftime('%Y-%m-%d %H:%M:%S')
    mask_vacio = df_participantes['Fecha log en app'].isna()
    df_participantes.loc[mask_vacio, 'Fecha log en app'] = match_fechas_str[mask_vacio].values
    
    # Asegurar columnas obligatorias
    columnas_obligatorias = [
        'Fecha log (solo día y mes)',
        'CUMPLE 4 SEMANAS EL',
        'CUMPLE LAS 6 SEMANAS EL',
        'Evaluable a las 4 semanas',
        'Evaluable a las 6 semanas',
        'Fecha último uso',
        'Última pantalla revisada',
        'Días desde último uso',
        'Fecha Esperada Finalización',
        'Fecha Finalización Real',
        'Pantalla Actual',
        'Completado',
        'Etapa del Embudo'
    ]
    for col in columnas_obligatorias:
        if col not in df_participantes.columns:
            df_participantes[col] = None
            
    if 'Completado' not in df_participantes.columns:
        df_participantes['Completado'] = False
    else:
        df_participantes['Completado'] = df_participantes['Completado'].astype(object)
    if 'Días Transcurridos' not in df_participantes.columns:
        df_participantes['Días Transcurridos'] = None
    if '% Progreso' not in df_participantes.columns:
        df_participantes['% Progreso'] = None
        
    for idx in df_participantes.index:
        tel = df_participantes.at[idx, 'tel_match']
        pantalla_actual_excel = df_participantes.at[idx, 'Pantalla Actual']
        
        # Filtro falsos positivos
        tiene_logs_servidor = tel in ultimas_fechas.index
        tiene_pantalla_excel = pd.notna(pantalla_actual_excel) and str(pantalla_actual_excel).strip() != "" and str(pantalla_actual_excel).strip().lower() != 'nan'
        
        if not tiene_logs_servidor and not tiene_pantalla_excel:
            df_participantes.at[idx, 'Fecha log en app'] = None
            df_participantes.at[idx, 'Fecha Esperada Finalización'] = None
            df_participantes.at[idx, 'Días Transcurridos'] = None
            df_participantes.at[idx, 'Fecha log (solo día y mes)'] = None
            df_participantes.at[idx, 'CUMPLE 4 SEMANAS EL'] = None
            df_participantes.at[idx, 'CUMPLE LAS 6 SEMANAS EL'] = None
            df_participantes.at[idx, 'Evaluable a las 4 semanas'] = None
            df_participantes.at[idx, 'Evaluable a las 6 semanas'] = None
            df_participantes.at[idx, 'Pantalla Actual'] = None
            df_participantes.at[idx, 'Última pantalla revisada'] = None
            df_participantes.at[idx, '% Progreso'] = 0.0
            df_participantes.at[idx, 'Fecha Última Actividad'] = None
            df_participantes.at[idx, 'Fecha último uso'] = None
            df_participantes.at[idx, 'Días desde último uso'] = None
            
        fecha_log = df_participantes.at[idx, 'Fecha log en app']
        
        if pd.notna(fecha_log):
            fecha_log_dt = parsear_timestamp(fecha_log)
            if pd.notna(fecha_log_dt):
                fecha_esperada = fecha_log_dt + timedelta(days=14)
                df_participantes.at[idx, 'Fecha Esperada Finalización'] = fecha_esperada
                df_participantes.at[idx, 'Días Transcurridos'] = (hoy_dt - fecha_log_dt).days
                df_participantes.at[idx, 'Fecha log (solo día y mes)'] = fecha_log_dt.strftime('%Y-%m-%d')
                
                date_log = fecha_log_dt.date()
                date_4w = date_log + timedelta(days=28)
                date_6w = date_log + timedelta(days=42)
                
                df_participantes.at[idx, 'CUMPLE 4 SEMANAS EL'] = date_4w.strftime('%Y-%m-%d')
                df_participantes.at[idx, 'CUMPLE LAS 6 SEMANAS EL'] = date_6w.strftime('%Y-%m-%d')
                
                if hoy_dt.date() >= date_4w:
                    df_participantes.at[idx, 'Evaluable a las 4 semanas'] = f"SÍ, DESDE EL {date_4w.strftime('%d/%m/%Y')}"
                else:
                    df_participantes.at[idx, 'Evaluable a las 4 semanas'] = f"NO, SE PODRÁ EL {date_4w.strftime('%d/%m/%Y')}"
                    
                if hoy_dt.date() >= date_6w:
                    df_participantes.at[idx, 'Evaluable a las 6 semanas'] = f"SÍ, DESDE EL {date_6w.strftime('%d/%m/%Y')}"
                else:
                    df_participantes.at[idx, 'Evaluable a las 6 semanas'] = f"NO, SE PODRÁ EL {date_6w.strftime('%d/%m/%Y')}"
                    
        pantalla = ultimas_pantallas.get(tel) if tel in ultimas_pantallas.index else None
        fecha_actividad = ultimas_fechas.get(tel) if tel in ultimas_fechas.index else None
        
        completado_original = df_participantes.at[idx, 'Completado']
        ya_completado = (pd.notna(completado_original) and (completado_original == True or completado_original == 1 or completado_original == 1.0 or str(completado_original).lower() == 'true'))
        
        if pd.notna(pantalla):
            df_participantes.at[idx, 'Pantalla Actual'] = pantalla
            df_participantes.at[idx, 'Última pantalla revisada'] = pantalla
            df_participantes.at[idx, '% Progreso'] = calcular_progreso(pantalla)
            
            if pd.notna(fecha_actividad):
                fecha_act_str = fecha_actividad.strftime('%Y-%m-%d %H:%M:%S')
                df_participantes.at[idx, 'Fecha Última Actividad'] = fecha_act_str
                df_participantes.at[idx, 'Fecha último uso'] = fecha_act_str
                df_participantes.at[idx, 'Días desde último uso'] = (hoy_dt - fecha_actividad).days
                
        estado_str = str(df_participantes.at[idx, 'Estado']).strip().lower()
        es_rechazo = any(p in estado_str for p in ['no tiene tiempo', 'no quieren', 'no quiere', 'no puede', 'rechaz', 'desist', 'baja', 'fuera']) or 'error' in estado_str
        es_falso_completado = 'falso' in estado_str or 'no completado' in estado_str
        
        if (tel in tels_con_final or ya_completado) and not es_rechazo and not es_falso_completado:
            df_participantes.at[idx, 'Completado'] = True
            df_participantes.at[idx, '% Progreso'] = 100.0
            fecha_final = fechas_final.get(tel, hoy_dt)
            if pd.notna(fecha_final):
                if isinstance(fecha_final, str):
                    df_participantes.at[idx, 'Fecha Finalización Real'] = fecha_final
                else:
                    df_participantes.at[idx, 'Fecha Finalización Real'] = fecha_final.strftime('%Y-%m-%d %H:%M:%S')
        elif es_falso_completado:
            df_participantes.at[idx, 'Completado'] = False
            df_participantes.at[idx, '% Progreso'] = 93.8
            df_participantes.at[idx, 'Fecha Finalización Real'] = None
            
    df_participantes = df_participantes.drop(columns=['tel_match'])
    
    # 3. Clasificación del Embudo
    def clasificar(row):
        estado = str(row.get('Estado', '')).strip().lower()
        fecha_log = row.get('Fecha log en app')
        pantalla = row.get('Pantalla Actual')
        completado = row.get('Completado', False)
        fecha_inscripcion = row.get('Fecha de inscripción en el servidor')
        
        if 'error' in estado:
            return 'Notificado con Error Técnico'
            
        patrones_rechazo = ['no tiene tiempo', 'no quieren', 'no quiere',
                            'no puede', 'rechaz', 'desist', 'baja', 'fuera']
        if any(p in estado for p in patrones_rechazo):
            return 'Rechazo / Sin disponibilidad'
            
        if pd.isna(fecha_log):
            if pd.notna(fecha_inscripcion):
                try:
                    fecha_inscripcion_dt = parsear_fecha_inscripcion(fecha_inscripcion)
                    if pd.notna(fecha_inscripcion_dt):
                        dias_trans = (hoy_dt - fecha_inscripcion_dt).days
                        if 0 <= dias_trans <= 5:
                            return 'Recién Notificado (< 5 días)'
                except Exception:
                    pass
            return 'Notificado - No Descarga App'
            
        if completado:
            return 'Completado'
            
        fecha_esperada = row.get('Fecha Esperada Finalización')
        if pd.notna(fecha_esperada):
            fecha_esp_dt = parsear_timestamp(fecha_esperada)
            if pd.notna(fecha_esp_dt):
                dias_restantes = (fecha_esp_dt - hoy_dt).days
                pct = calcular_progreso(pantalla) if pd.notna(pantalla) else 0
                
                if pct >= 75:
                    return 'En Progreso - Próximo a Terminar (Avanzado)'
                elif pct >= 50:
                    return 'En Progreso - Empezando Libro 2'
                elif dias_restantes < 0:
                    return 'En Progreso - Atrasado'
                else:
                    return 'En Libro 1'
                    
        return 'En Libro 1'
        
    df_participantes['Status Clasificado'] = df_participantes.apply(clasificar, axis=1)
    df_participantes['Status Ordenado'] = df_participantes['Status Clasificado'].map(status_map)
    
    return df_participantes, nuevos_pendientes

# ==============================================================================
# 💾 EXPORTACIÓN PRESERVANDO ESTILOS
# ==============================================================================
def guardar_archivo_final(ruta_plantilla, ruta_salida, df_datos, df_original=None, sheet_name="PARTICIPANTES 2", keep_only_target_sheet=True):
    """
    Toma el archivo original subido, sobreescribe la hoja especificada
    aplicando colores a los cambios y fila por fila. Si keep_only_target_sheet es True,
    deja únicamente la hoja de participantes en el Excel final.
    """
    from openpyxl.styles import PatternFill
    
    shutil.copy(ruta_plantilla, ruta_salida)
    wb = openpyxl.load_workbook(ruta_salida)
    nombre_hoja = sheet_name
    
    if nombre_hoja not in wb.sheetnames:
        ws = wb.create_sheet(title=nombre_hoja)
        start_row = 1
    else:
        ws = wb[nombre_hoja]
        start_row = 3  # Fila 3 cabecera
        
    # Escribir cabeceras
    for c_idx, col_name in enumerate(df_datos.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=col_name)
        
    fill_verde_nuevo = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_amarillo_cambio = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_naranja_celda = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    orig_dict = {}
    if df_original is not None:
        try:
            id_col_orig = df_original.columns[0]
            for r_idx_orig, row_orig in df_original.iterrows():
                pid = str(row_orig[id_col_orig]).strip().replace('.0','')
                orig_dict[pid] = row_orig.to_dict()
        except Exception as e:
            print(f"⚠️ Error comparando originales para excel: {e}")
            
    num_filas_df = len(df_datos)
    id_col_datos = df_datos.columns[0]
    
    for r_offset in range(num_filas_df):
        r_idx = start_row + 1 + r_offset
        pid = str(df_datos.iloc[r_offset][id_col_datos]).strip().replace('.0','')
        
        is_new = pid not in orig_dict
        celdas_cambiadas = []
        
        if not is_new and df_original is not None:
            orig_row = orig_dict[pid]
            for c_idx, col_name in enumerate(df_datos.columns, 1):
                if col_name in orig_row:
                    val_nuevo = df_datos.iloc[r_offset, c_idx - 1]
                    val_viejo = orig_row[col_name]
                    if valores_diferentes(val_nuevo, val_viejo):
                        celdas_cambiadas.append(c_idx)
                        
        for c_idx, col_name in enumerate(df_datos.columns, 1):
            val = df_datos.iloc[r_offset, c_idx - 1]
            
            if pd.isna(val) or (isinstance(val, float) and np.isnan(val)) or val == 'NaT' or val == 'nan':
                val_formatted = None
            elif isinstance(val, (datetime, pd.Timestamp)):
                if val.hour == 0 and val.minute == 0 and val.second == 0:
                    val_formatted = val.strftime('%Y-%m-%d')
                else:
                    val_formatted = val.strftime('%Y-%m-%d %H:%M:%S')
            else:
                if col_name == 'Completado':
                    val_formatted = bool(val) if pd.notna(val) else False
                else:
                    val_formatted = val
                    
            cell = ws.cell(row=r_idx, column=c_idx, value=val_formatted)
            
            # Aplicar colores
            if is_new:
                cell.fill = fill_verde_nuevo
            elif len(celdas_cambiadas) > 0:
                if c_idx in celdas_cambiadas:
                    cell.fill = fill_naranja_celda
                else:
                    cell.fill = fill_amarillo_cambio
                    
    max_row_original = ws.max_row
    last_written_row = start_row + num_filas_df
    if max_row_original > last_written_row:
        ws.delete_rows(last_written_row + 1, max_row_original - last_written_row)
        
    # Eliminar otras pestañas si se solicita
    if keep_only_target_sheet:
        for s_name in list(wb.sheetnames):
            if s_name != nombre_hoja:
                try:
                    wb.remove(wb[s_name])
                except Exception as e:
                    print(f"⚠️ Error al eliminar pestaña {s_name}: {e}")
                    
    wb.save(ruta_salida)
